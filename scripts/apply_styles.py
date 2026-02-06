import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Force UTF-8 for Windows console
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass

def get_technical_notes_severity(notes: str) -> str:
    """
    Анализирует Technical Notes и возвращает severity level.
    """
    if not notes or str(notes).strip() == "—":
        return 'OK'
    
    notes_lower = str(notes).lower()
    
    # CRITICAL markers
    critical_markers = [
        '❌ http only',
        '⏱️ site unreachable',
        '🔥 http',
        '🎨 adobe flash'
    ]
    
    # HIGH markers
    high_markers = [
        '⚠️ ssl',
        '📱 missing',
        '📱 no responsive',
        '🎨 jquery'
    ]
    
    # MEDIUM markers
    medium_markers = [
        '⚡ slow',
        '⚡ very slow',
        '🎨 table-based',
        '🎨 fixed-width',
        '🎨 outdated',
        '⚠️ outdated'
    ]
    
    # Check in priority order
    for marker in critical_markers:
        if marker in notes_lower:
            return 'CRITICAL'
    
    for marker in high_markers:
        if marker in notes_lower:
            return 'HIGH'
    
    for marker in medium_markers:
        if marker in notes_lower:
            return 'MEDIUM'
    
    if '✅' in str(notes):
        return 'OK'
    
    return 'MEDIUM'  # default

def apply_excel_formatting(filepath):
    """
    Применяет визуальное форматирование к Excel файлу.
    """
    print(f"🎨 Применяю стили к: {filepath}")
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        
        # Цвета для заливки
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        orange_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")
        green_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
        gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Шрифты
        header_font = Font(bold=True, color="FFFFFF", size=12)
        bold_font = Font(bold=True)
        
        # Выравнивание
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        top_left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Для Technical Notes
        
        # Границы
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        # 1. Форматирование заголовков (первая строка)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = thin_border
        
        ws.row_dimensions[1].height = 30
        
        # Закрепляем первую строку (заголовок)
        ws.freeze_panes = 'A2'
        
        # 2. Автоматическое растягивание колонок по содержимому
        for col_idx, column in enumerate(ws.columns, 1):
            col_letter = get_column_letter(col_idx)
            
            # Получаем название колонки
            header_cell = ws.cell(1, col_idx)
            column_name = header_cell.value if header_cell.value else ""
            
            # Находим максимальную длину содержимого в колонке
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        cell_value = str(cell.value)
                        cell_length = len(cell_value)
                        # Для эмодзи добавляем дополнительное пространство
                        emoji_count = sum(1 for char in cell_value if ord(char) > 0x1F300)
                        cell_length += emoji_count
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Устанавливаем ширину с учетом типа колонки
            if column_name == '💬 Technical Notes':
                adjusted_width = 60
            elif column_name in ['Category', 'Name', 'Address', 'Website', 'Email']:
                max_width = 70
                adjusted_width = max(8, min(max_length + 3, max_width))
            else:
                max_width = 50
                adjusted_width = max(8, min(max_length + 3, max_width))
            
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # 3. Форматирование ячеек данных + цветовое кодирование
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row, start=1):
                cell.border = thin_border
                cell.alignment = left_alignment
                
                # Получаем название колонки через заголовок
                column_name = ws.cell(1, col_idx).value
                
                # Technical Notes - специальное форматирование
                if column_name == '💬 Technical Notes':
                    cell.alignment = top_left_alignment
                    ws.row_dimensions[row_idx].height = None
                    
                    notes_text = str(cell.value) if cell.value else ""
                    severity = get_technical_notes_severity(notes_text)
                    
                    if severity == 'CRITICAL':
                        cell.fill = red_fill
                    elif severity == 'HIGH':
                        cell.fill = orange_fill
                    elif severity == 'MEDIUM':
                        cell.fill = yellow_fill
                    elif severity == 'OK':
                        cell.fill = green_fill
                    
                    continue
                
                # Цветовое кодирование по содержимому
                value = str(cell.value) if cell.value else ""
                
                if column_name == '#' or column_name == '⭐ Customer Rating' or column_name == 'Phone':
                    cell.alignment = center_alignment
                
                if "✅ Online" in value:
                    cell.fill = green_fill
                elif "⚠️ Protected" in value:
                    cell.fill = orange_fill
                elif "❌ Offline" in value or "⏱️ Timeout" in value:
                    cell.fill = red_fill
                elif "✅ HTTPS + SSL" in value:
                    cell.fill = green_fill
                elif "⚠️ HTTPS" in value:
                    cell.fill = orange_fill
                elif "❌ HTTP Only" in value:
                    cell.fill = red_fill
                elif "🟢 Fast" in value:
                    cell.fill = green_fill
                elif "🟡 Average" in value:
                    cell.fill = yellow_fill
                elif "🔴 Slow" in value:
                    cell.fill = red_fill
                elif "📱 Mobile" in str(column_name):
                    if "✅ Yes" in value:
                        cell.fill = green_fill
                    elif "❌ No" in value:
                        cell.fill = red_fill
                elif "🎨 Design" in str(column_name):
                    if "✅ Modern" in value:
                        cell.fill = green_fill
                    elif "⚠️ Outdated" in value:
                        cell.fill = orange_fill
                elif "🔍 SEO" in str(column_name):
                    if "✅ Optimized" in value:
                        cell.fill = green_fill
                    elif "⚠️ Basic" in value:
                        cell.fill = orange_fill
                    elif "❌ Missing" in value:
                        cell.fill = red_fill
                elif column_name == '⭐ Customer Rating' and value != "—":
                    try:
                        rating_val = float(value)
                        if rating_val >= 4.5:
                            cell.fill = green_fill
                        elif rating_val >= 4.0:
                            cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                        elif rating_val >= 3.5:
                            cell.fill = yellow_fill
                        elif rating_val >= 3.0:
                            cell.fill = orange_fill
                        else:
                            cell.fill = red_fill
                    except:
                        pass
                elif "🔥 HOT LEAD" in value:
                    cell.fill = red_fill
                    cell.font = bold_font
                elif "🟠 QUALIFIED" in value:
                    cell.fill = orange_fill
                elif "🟡 POTENTIAL" in value:
                    cell.fill = yellow_fill
                elif "⚪ SKIP" in value:
                    cell.fill = gray_fill
        
        # 4. Включаем автофильтры
        if ws.max_row > 0:
            ws.auto_filter.ref = ws.dimensions
        
        wb.save(filepath)
        print("✅ Успешно применены стили!")
        
    except Exception as e:
        print(f"❌ Ошибка форматирования: {e}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Использование: python apply_styles.py <file.xlsx>")
    else:
        apply_excel_formatting(sys.argv[1])
