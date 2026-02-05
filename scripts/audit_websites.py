#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Website Audit Script for BrokenSite Hunter
===========================================

Автоматически проверяет сайты на:
- SSL валидность (просроченные сертификаты)
- PageSpeed score (скорость загрузки)
- HTTPS vs HTTP
- Доступность сайта

Использование:
    python audit_websites.py input.csv output.csv
"""

import ssl
import socket
import requests
import pandas as pd
import time
import re
from urllib.parse import urlparse
from datetime import datetime
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Import helper functions for v2.3
from audit_websites_helpers import (
    check_https_multistep,
    check_mobile_multistep,
    extract_jquery_version,
    get_jquery_release_year,
    check_seo,
    extract_emails_from_website
)

# Import deduplication functions for v2.4
from audit_deduplication import deduplicate_data

# ============================================================================
# НАСТРОЙКА КОДИРОВКИ ДЛЯ WINDOWS
# ============================================================================

# Устанавливаем UTF-8 для вывода в консоль (для emoji)
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        # Для более старых версий Python
        import codecs
        sys.stdout = codecs.getwriter('utf-8')(sys.stdout.buffer, 'strict')
        sys.stderr = codecs.getwriter('utf-8')(sys.stderr.buffer, 'strict')


# ============================================================================
# КОНФИГУРАЦИЯ
# ============================================================================

# Google PageSpeed API (бесплатно 25,000 запросов/день)
# Получить ключ: https://console.cloud.google.com/apis/credentials
PAGESPEED_API_KEY = os.getenv('PAGESPEED_API_KEY', '')  # Опционально

# Rate limiting (чтобы не заблокировали)
DELAY_BETWEEN_REQUESTS = 2  # секунд между проверками
REQUEST_TIMEOUT = 15  # секунд на запрос (увеличено для медленных сайтов)

# ============================================================================
# ФУНКЦИИ ПРОВЕРКИ
# ============================================================================

def normalize_url(url):
    """
    Нормализует URL для проверки.
    
    Примеры:
        example.com -> https://example.com
        www.example.com -> https://www.example.com
        http://example.com -> http://example.com
    """
    if not url or pd.isna(url):
        return None
    
    url = str(url).strip()
    
    # Убираем пробелы и лишние символы
    url = url.replace(' ', '')
    
    # Если уже есть протокол - оставляем как есть
    if url.startswith('http://') or url.startswith('https://'):
        return url
    
    # Добавляем https:// по умолчанию
    return f'https://{url}'


def check_ssl(url):
    """
    Проверяет валидность SSL сертификата.
    
    Returns:
        dict: {
            'valid': bool,
            'error': str or None
        }
    """
    try:
        hostname = urlparse(url).hostname
        if not hostname:
            return {'valid': False, 'error': 'Invalid hostname'}
        
        context = ssl.create_default_context()
        
        with socket.create_connection((hostname, 443), timeout=5) as sock:
            with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                cert = ssock.getpeercert()
                # Сертификат получен = валидный
                return {'valid': True, 'error': None}
                
    except ssl.SSLError as e:
        return {'valid': False, 'error': f'SSL Error: {str(e)[:50]}'}
    except socket.timeout:
        return {'valid': False, 'error': 'Timeout'}
    except Exception as e:
        return {'valid': False, 'error': f'Error: {str(e)[:50]}'}


def check_https(url):
    """
    Проверяет, использует ли сайт HTTPS.
    
    Returns:
        bool: True если HTTPS, False если HTTP
    """
    return url.startswith('https://')


def check_site_availability(url):
    """
    Проверяет, доступен ли сайт (возвращает 200 OK).
    Также измеряет время загрузки в секундах.
    
    Returns:
        dict: {
            'available': bool,
            'status_code': int or None,
            'error': str or None,
            'load_time': float (секунды)
        }
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
        }
        
        # Засекаем время начала
        start_time = time.time()
        
        response = requests.get(
            url,
            headers=headers,
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,
            verify=False  # Игнорируем SSL ошибки для этой проверки
        )
        
        # Засекаем время окончания
        load_time = time.time() - start_time
        
        if response.status_code == 200:
            return {
                'available': True,
                'status_code': 200,
                'error': None,
                'load_time': load_time
            }
        else:
            return {
                'available': False,
                'status_code': response.status_code,
                'error': f'HTTP {response.status_code}',
                'load_time': load_time
            }
            
    except requests.exceptions.Timeout:
        return {'available': False, 'status_code': None, 'error': 'Timeout', 'load_time': None}
    except requests.exceptions.ConnectionError:
        return {'available': False, 'status_code': None, 'error': 'Connection Error', 'load_time': None}
    except Exception as e:
        return {'available': False, 'status_code': None, 'error': str(e)[:50], 'load_time': None}


def check_mobile_friendly(html_content):
    """
    Проверяет наличие viewport meta tag (признак мобильной версии).
    Улучшенная версия: ищем разные варианты viewport.
    
    Returns:
        dict: {
            'is_mobile_friendly': bool,
            'has_viewport': bool
        }
    """
    if not html_content:
        return {'is_mobile_friendly': False, 'has_viewport': False}
    
    html_lower = html_content.lower()
    
    # Проверяем разные варианты viewport meta tag
    has_viewport = False
    
    # Вариант 1: Стандартный viewport (самый распространенный)
    if 'name="viewport"' in html_lower or "name='viewport'" in html_lower:
        has_viewport = True
    
    # Вариант 2: Проверяем content="width=device-width"
    if 'width=device-width' in html_lower:
        has_viewport = True
    
    # Вариант 3: viewport с initial-scale
    if 'initial-scale' in html_lower and 'viewport' in html_lower:
        has_viewport = True
    
    # Дополнительно: Проверяем наличие media queries (признак адаптивности)
    has_media_queries = '@media' in html_lower or 'media="screen' in html_lower
    
    # Считаем мобильным если есть viewport ИЛИ есть media queries
    is_mobile_friendly = has_viewport or has_media_queries
    
    return {
        'is_mobile_friendly': is_mobile_friendly,
        'has_viewport': has_viewport
    }



def check_outdated_design(html_content, url):
    """
    Проверяет признаки устаревшего дизайна.
    
    Returns:
        dict: {
            'is_outdated': bool,
            'issues': list[str]  # Список найденных проблем
        }
    """
    if not html_content:
        return {'is_outdated': False, 'issues': []}
    
    html_lower = html_content.lower()
    issues = []
    
    # 1. Использование таблиц для layout (старая практика)
    if '<table' in html_lower and 'cellpadding' in html_lower:
        issues.append("Table layout")
    
    # 2. Flash (устарел полностью)
    if 'flash' in html_lower or '.swf' in html_lower:
        issues.append("Flash detected")
    
    # 3. Старые jQuery версии (1.x)
    if 'jquery-1.' in html_lower or 'jquery/1.' in html_lower:
        issues.append("Old jQuery")
    
    # 4. Фиксированная ширина (не responsive)
    if 'width:' in html_lower and 'px' in html_lower and '@media' not in html_lower:
        issues.append("Fixed width")
    
    return {
        'is_outdated': len(issues) > 0,
        'issues': issues
    }


def check_pagespeed(url, api_key=None):
    """
    Проверяет PageSpeed score через Google API.
    
    Returns:
        dict: {
            'score': int (0-100) or None,
            'error': str or None
        }
    """
    if not api_key:
        # Без API ключа используем публичный эндпоинт (лимитированный)
        print(f"  ⚠️  PageSpeed: Нет API ключа, пропускаем...")
        return {'score': None, 'error': 'No API key'}
    
    try:
        api_url = f"https://www.googleapis.com/pagespeedonline/v5/runPagespeed"
        params = {
            'url': url,
            'key': api_key,
            'category': 'PERFORMANCE'
        }
        
        response = requests.get(api_url, params=params, timeout=30)
        
        if response.status_code == 200:
            data = response.json()
            score = data['lighthouseResult']['categories']['performance']['score']
            score = int(score * 100)  # Конвертируем 0.XX в 0-100
            return {'score': score, 'error': None}
        else:
            return {'score': None, 'error': f'API Error: {response.status_code}'}
            
    except Exception as e:
        return {'score': None, 'error': str(e)[:50]}


def generate_technical_notes(audit_result: dict) -> str:
    """
    Generates concise, personalized technical comments (3-5 lines max).
    Output in English only.
    
    Returns:
        str: Brief technical notes with line breaks (\n) and trailing newline
    """
    notes = []
    
    # Priority 1: Site Availability Issues
    if not audit_result.get('site_available', True):
        error = audit_result.get('availability_error', 'Unknown')
        status_code = audit_result.get('status_code')
        
        if 'Timeout' in error:
            notes.append(f"⏱️ Site unreachable: Server timeout (>{REQUEST_TIMEOUT}s)")
            notes.append("Check hosting or server overload")
            return "\n".join(notes) + "\n"
        
        elif status_code == 403 or 'Protected (403)' in error:
            notes.append("⚠️ HTTP 403: Anti-bot protection active (Cloudflare/WAF)")
            notes.append("Site works for real users - not critical")
            return "\n".join(notes) + "\n"
        
        elif status_code == 404:
            notes.append("❌ HTTP 404: Page not found")
            notes.append("Domain may be deleted or URL changed")
            return "\n".join(notes) + "\n"
        
        elif status_code and status_code >= 500:
            notes.append(f"🔥 HTTP {status_code}: Server error - site completely down!")
            notes.append("HOT LEAD - requires immediate attention")
            return "\n".join(notes) + "\n"
    
    # Priority 2: Security Issues (HTTPS/SSL)
    is_https = audit_result.get('is_https', False)
    ssl_valid = audit_result.get('ssl_valid', False)
    ssl_error = audit_result.get('ssl_error', '')
    
    if not is_https:
        notes.append("❌ HTTP only (no encryption)")
        notes.append("Google marks as 'Not Secure' - critical security issue")
    elif is_https and not ssl_valid and ssl_error != 'Not HTTPS':
        if ssl_error and 'expired' in ssl_error.lower():
            notes.append("⚠️ SSL Certificate Expired (Security Risk)")
        elif ssl_error and 'self-signed' in ssl_error.lower():
            notes.append("⚠️ Self-Signed Certificate")
        else:
            notes.append(f"❌ SSL Invalid: {ssl_error or 'Unknown Error'}")
    
    # Priority 3: Mobile Issues (specific details)
    is_mobile = audit_result.get('is_mobile_friendly', True)
    mobile_check = audit_result.get('mobile_check', {})
    
    if not is_mobile:
        reasons = mobile_check.get('reasons', [])
        
        # Find specific issue
        if any('viewport' in r.lower() for r in reasons):
            notes.append("📱 Missing <meta name='viewport'> tag")
            notes.append("Site doesn't adapt to phone screens")
        elif any('media' in r.lower() for r in reasons):
            notes.append("📱 No responsive CSS (@media queries)")
            notes.append("60%+ visitors on mobile see broken layout")
        else:
            notes.append("📱 Not mobile-optimized")
            notes.append("Poor UX on phones - users leave quickly")
    
    # Priority 4: Design Issues (specific technologies)
    is_outdated = audit_result.get('is_outdated', False)
    design_check = audit_result.get('design_check', {})
    
    if is_outdated:
        issues = design_check.get('issues', [])
        jquery_version = design_check.get('jquery_version')
        
        # Prioritize critical issues with specific details
        if 'Flash detected' in issues:
            notes.append("🎨 Adobe Flash detected")
            notes.append("Component broken - Flash died in 2020")
        elif 'Old jQuery' in issues and jquery_version:
            year = get_jquery_release_year(jquery_version)
            notes.append(f"🎨 jQuery {jquery_version} from {year}")
            notes.append("Security vulnerabilities - needs update")
        elif 'Table layout' in issues:
            notes.append("🎨 Table-based layout (1990s tech)")
            notes.append("Modern sites use CSS Grid/Flexbox")
        elif 'Fixed width' in issues:
            notes.append("🎨 Fixed-width design (not responsive)")
            notes.append("Doesn't resize for different screens")

    # Priority 5: SEO Issues (v2.6)
    seo_status = audit_result.get('seo_status', '✅ Optimized')
    seo_details = audit_result.get('seo_details', '')
    
    if 'Missing' in seo_status:
        notes.append("🔍 Critical SEO: Missing tags")
        notes.append(f"{seo_details}") # e.g. "Missing Title; Missing Desc"
    elif 'Basic' in seo_status and seo_details != "All Good":
        # Limit details length
        details_short = seo_details.split(';')[0]
        notes.append(f"🔍 SEO Improvement needed")
        notes.append(f"{details_short}")
    
    # Priority 5: Speed Issues (specific metrics)
    load_time = audit_result.get('load_time')
    if load_time and load_time > 3.0:
        if load_time > 5.0:
            notes.append(f"⚡ Very slow: {load_time:.1f}s load time")
            notes.append("50% of users abandon after 3s")
        else:
            notes.append(f"⚡ Slow: {load_time:.1f}s (should be <3s)")
            notes.append("Hurts Google rankings & user retention")
    
    # Return
    if not notes:
        return "✅ All checks passed\n"
    
    # Limit to 5 lines max + add trailing newline for visual spacing
    return "\n".join(notes[:5]) + "\n"



def audit_website(url, api_key=None, index=0, total=0):
    """
    Полная проверка сайта.
    
    Returns:
        dict: Результаты всех проверок
    """
    print(f"\n[{index}/{total}] Проверяю: {url}")
    
    normalized_url = normalize_url(url)
    
    if not normalized_url:
        print("  ❌ Invalid URL")
        return {
            'url_normalized': None,
            'final_url': None,
            'is_https': False,
            'site_available': False,
            'status_code': None,
            'ssl_valid': False,
            'ssl_error': 'Invalid URL',
            'pagespeed_score': None,
            'issue_severity': 'UNKNOWN'
        }
    
    # =========================================================================
    # КРИТИЧЕСКОЕ ИСПРАВЛЕНИЕ: Делаем ОДИН запрос с получением финального URL
    # =========================================================================
    
    html_content = None
    final_url = normalized_url
    load_time = None
    status_code = None
    site_available = False
    availability_error = None
    
    try:
        # Более реалистичный User-Agent для обхода антибот защиты
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'DNT': '1',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1'
        }
        
        # Засекаем время начала
        start_time = time.time()
        
        # Делаем запрос с редиректами
        response = requests.get(
            normalized_url,
            headers=headers,
            timeout=REQUEST_TIMEOUT,
            allow_redirects=True,  # Следуем редиректам
            verify=False  # Игнорируем SSL ошибки для этой проверки
        )
        
        # Засекаем время окончания
        load_time = time.time() - start_time
        
        # Сохраняем финальный URL после редиректов
        final_url = response.url
        status_code = response.status_code
        
        # Считаем сайт доступным если:
        # - 2xx (успех)
        # - 403 (работает, но блокирует бота - это НЕ критично)
        # Коды 3xx уже обработаны через allow_redirects=True
        if 200 <= response.status_code < 300:
            site_available = True
            html_content = response.text
        elif response.status_code == 403:
            # Сайт работает, но блокирует автоматические запросы
            site_available = True  # Считаем доступным!
            html_content = response.text if response.text else None
            availability_error = 'Protected (403)'  # Помечаем как защищенный
        else:
            # Для остальных кодов (404, 500, etc.) - сайт недоступен
            availability_error = f'HTTP {response.status_code}'
            
    except requests.exceptions.Timeout:
        availability_error = 'Timeout'
    except requests.exceptions.ConnectionError:
        availability_error = 'Connection Error'
    except Exception as e:
        availability_error = str(e)[:50]
    
    # =========================================================================
    # V2.3: МНОГОСТУПЕНЧАТАЯ ПРОВЕРКА
    # =========================================================================
    
    # Get response object for headers
    response_headers = response.headers if 'response' in locals() and response else None
    
    # 1. HTTPS + SSL Check (Multi-step с внешними API если нужно)
    https_check = check_https_multistep(
        url=final_url,
        html_content=html_content,
        response_headers=response_headers,
        response_obj=response if 'response' in locals() else None
    )
    
    is_https = https_check.get('is_https', False)
    ssl_valid = https_check.get('ssl_valid', False)
    ssl_error = https_check.get('ssl_error', 'Not HTTPS')
    
    print(f"  🔒 HTTPS: {'✅ Yes' if is_https else '❌ No (HTTP)'}")
    if is_https:
        ssl_status = '✅ Valid' if ssl_valid else f"❌ Invalid ({ssl_error})"
        print(f"  🔐 SSL: {ssl_status}")
        
        # Показываем методы проверки
        methods = https_check.get('methods', [])
        if len(methods) > 1:
            print(f"     Проверено методами: {', '.join(methods)}")
    
    # 2. Site Availability
    avail_status = '✅ Yes' if site_available else f"❌ No ({availability_error})"
    print(f"  🌐 Available: {avail_status}")
    
    # Если был редирект - показываем финальный URL
    if final_url != normalized_url:
        print(f"  🔄 Redirected to: {final_url}")
    
    # 3. Mobile Friendly Check (Multi-step, только если HTML доступен)
    mobile_check = {'is_mobile_friendly': False, 'reasons': [], 'methods': []}
    if html_content:
        mobile_check = check_mobile_multistep(html_content, final_url)
        is_mobile_friendly = mobile_check.get('is_mobile_friendly', False)
        mobile_status = '✅ Yes' if is_mobile_friendly else '❌ No'
        print(f"  📱 Mobile: {mobile_status}")
        
        # Показываем методы проверки
        mobile_methods = mobile_check.get('methods', [])
        if len(mobile_methods) > 1:
            print(f"     Проверено методами: {', '.join(mobile_methods[:2])}")
    else:
        is_mobile_friendly = False
    
    # 4. Outdated Design Check (с извлечением jQuery версии)
    design_check = {'is_outdated': False, 'issues': [], 'jquery_version': None}
    if html_content:
        design_check = check_outdated_design(html_content, final_url)
        
        # Извлекаем версию jQuery если найден
        if 'Old jQuery' in design_check.get('issues', []):
            jquery_version = extract_jquery_version(html_content)
            design_check['jquery_version'] = jquery_version
        
        if design_check.get('is_outdated'):
            # Получаем список issues - совместимость со старой версией
            issues_list = design_check.get('issues', design_check.get('reasons', []))
            print(f"  🎨 Design: ⚠️ Outdated ({', '.join(issues_list[:2])}")
    
    is_outdated = design_check.get('is_outdated', False)
    
    # 5. SEO Audit (v2.6)
    seo_result = {'status': '❌ Missing', 'details': '', 'score': 0}
    if html_content:
        seo_result = check_seo(html_content)
        print(f"  🔍 SEO: {seo_result['status']}")
        if seo_result.get('details') and seo_result['details'] != "All Good":
             print(f"     Issues: {seo_result['details']}")
    
    # 6. Email Extraction (v2.7)
    scraped_emails = []
    if site_available and html_content:
        print("  📧 Extracting emails...")
        scraped_emails = extract_emails_from_website(final_url, html_content)
        if scraped_emails:
            print(f"  📧 Found: {', '.join(scraped_emails)}")
        else:
            print("  📧 No emails found")
    
    # 5. PageSpeed Check (только если сайт доступен)
    pagespeed_result = {'score': None, 'error': 'Site unavailable'}
    if site_available and api_key:
        time.sleep(1)  # Задержка перед API запросом
        pagespeed_result = check_pagespeed(final_url, api_key)
        if pagespeed_result['score'] is not None:
            print(f"  ⚡ PageSpeed: {pagespeed_result['score']}/100")
    
    # Определение серьезности проблемы
    severity = determine_severity(
        is_https=is_https,
        ssl_valid=ssl_valid,
        available=site_available,
        pagespeed=pagespeed_result['score'],
        is_mobile_friendly=is_mobile_friendly,
        is_outdated=is_outdated
    )
    
    # Формируем результат аудита для Technical Notes
    audit_result_for_notes = {
        'site_available': site_available,
        'availability_error': availability_error,
        'status_code': status_code,
        'is_https': is_https,
        'ssl_valid': ssl_valid,
        'ssl_error': ssl_error,
        'https_check': https_check,
        'is_mobile_friendly': is_mobile_friendly,
        'mobile_check': mobile_check,
        'is_outdated': is_outdated,
        'design_check': design_check,
        'seo_status': seo_result['status'],   # Flattened for generate_technical_notes
        'seo_details': seo_result['details'], # Flattened for generate_technical_notes
        'load_time': load_time
    }
    
    # Генерируем Technical Notes
    technical_notes = generate_technical_notes(audit_result_for_notes)
    
    return {
        'url_normalized': normalized_url,
        'final_url': final_url,
        'is_https': is_https,
        'site_available': site_available,
        'availability_error': availability_error,
        'status_code': status_code,
        'load_time': load_time,
        'ssl_valid': ssl_valid,
        'ssl_error': ssl_error,
        'is_mobile_friendly': is_mobile_friendly,
        'is_outdated': is_outdated,
        'design_issues': ', '.join(design_check.get('issues', [])) if design_check.get('issues') else None,
        'seo_status': seo_result['status'],
        'pagespeed_score': pagespeed_result['score'],
        'issue_severity': severity,
        'technical_notes': technical_notes,  # НОВОЕ ПОЛЕ v2.3!
        'scraped_emails': ', '.join(scraped_emails) if scraped_emails else None # v2.7
    }


def determine_severity(is_https, ssl_valid, available, pagespeed, is_mobile_friendly=True, is_outdated=False):
    """
    Определяет серьезность проблем на сайте.
    Теперь учитывает: HTTPS, SSL, доступность, скорость, мобильную версию и дизайн.
    
    Returns:
        str: 'CRITICAL', 'HIGH', 'MEDIUM', 'LOW', 'OK'
    """
    if not available:
        return 'CRITICAL'  # Сайт вообще не работает
    
    if not is_https:
        return 'CRITICAL'  # HTTP в 2026 = критично
    
    if is_https and not ssl_valid:
        return 'CRITICAL'  # Просроченный SSL = клиенты видят предупреждение
    
    # HIGH - нет мобильной версии (40%+ пользователей на мобильных)
    if not is_mobile_friendly:
        return 'HIGH'
    
    # HIGH - устаревший дизайн (отпугивает клиентов)
    if is_outdated:
        return 'HIGH'
    
    # Проверка PageSpeed
    if pagespeed is not None:
        if pagespeed < 30:
            return 'HIGH'  # Очень медленный сайт
        elif pagespeed < 50:
            return 'MEDIUM'  # Медленный сайт
        elif pagespeed < 70:
            return 'LOW'  # Можно улучшить
    
    return 'OK'  # Все хорошо


# ============================================================================
# ФУНКЦИИ ВИЗУАЛИЗАЦИИ
# ============================================================================

def format_website_status(available, error=None):
    """
    Форматирует статус доступности сайта с эмодзи.
    
    Returns:
        str: "✅ Online" / "⚠️ Protected" / "❌ Offline" / "⏱️ Timeout"
    """
    if available and error and 'Protected' in str(error):
        return "⚠️ Protected"  # Работает, но блокирует ботов
    elif available:
        return "✅ Online"
    elif error and 'timeout' in str(error).lower():
        return "⏱️ Timeout"
    else:
        return "❌ Offline"


def format_security_status(is_https, ssl_valid, ssl_error=None):
    """
    Форматирует статус безопасности с эмодзи.
    
    Returns:
        str: "✅ HTTPS + SSL" / "⚠️ HTTPS (SSL Error)" / "❌ HTTP Only"
    """
    if is_https and ssl_valid:
        return "✅ HTTPS + SSL"
    elif is_https and not ssl_valid:
        return "⚠️ HTTPS (SSL Error)"
    else:
        return "❌ HTTP Only"


def clean_website_url(url):
    """
    Форматирует URL для Excel: добавляет www., оставляет только домен.
    Excel распознает такие URL как кликабельные ссылки.
    
    Пример:
        http://www.mintdds.com/ → www.mintdds.com
        https://bostondental.com/boston-downtown-crossing/ → www.bostondental.com
        dentologyboston.com/index.html → www.dentologyboston.com
    
    Returns:
        str: URL в формате www.domain.com (без путей)
    """
    if not url or pd.isna(url):
        return ""
    
    url = str(url).strip()
    
    # Убираем протокол
    url = url.replace('https://', '').replace('http://', '')
    
    # Убираем слэш в конце
    if url.endswith('/'):
        url = url[:-1]
    
    # Извлекаем только домен (убираем пути типа /boston-downtown-crossing/)
    # Разделяем по "/" и берем первую часть
    url = url.split('/')[0]
    
    # Убираем query параметры и якоря
    url = url.split('?')[0].split('#')[0]
    
    # Добавляем www. если его нет
    if not url.startswith('www.'):
        url = 'www.' + url
    
    return url


def format_rating(rating):
    """
    Форматирует рейтинг БЕЗ звезды (звезда теперь в названии колонки).
    
    Returns:
        str: "4.5" или "—"
    """
    if rating is None or pd.isna(rating):
        return "—"
    
    try:
        rating_val = float(rating)
        return f"{rating_val}"
    except:
        return "—"


def format_emails(emails):
    """
    Форматирует email адреса: добавляет пробел после запятой.
    
    Пример:
        "email1@test.com,email2@test.com" → "email1@test.com, email2@test.com"
    
    Returns:
        str: Форматированный список email
    """
    if not emails or pd.isna(emails):
        return "—"
    
    emails_str = str(emails).strip()
    
    # Filter out placeholders like "### In progress ###"
    if "progress" in emails_str.lower() or "###" in emails_str:
        return "—"

    # Добавляем пробел после запятой если его нет
    if ',' in emails_str:
        # Разделяем по запятой, убираем пробелы, соединяем с ", "
        email_list = [e.strip() for e in emails_str.split(',')]
        emails_str = ', '.join(email_list)
    
    return emails_str


def format_speed_score(load_time):
    """
    Форматирует время загрузки с эмодзи.
    
    Args:
        load_time: Время загрузки в секундах
    
    Returns:
        str: "0.5s 🟢 Fast" / "2.3s 🟡 Average" / "5.1s 🔴 Slow" / "—"
    """
    if load_time is None or pd.isna(load_time):
        return "—"
    
    try:
        load_time = float(load_time)
        
        # Категории скорости
        if load_time < 1.0:
            return f"{load_time:.2f}s 🟢 Fast"
        elif load_time < 3.0:
            return f"{load_time:.2f}s 🟡 Average"
        else:
            return f"{load_time:.2f}s 🔴 Slow"
    except:
        return "—"


def determine_lead_quality(severity, rating):
    """
    Определяет качество лида на основе severity и rating.
    
    Returns:
        str: "🔥 HOT LEAD" / "🟠 QUALIFIED" / "🟡 POTENTIAL" / "⚪ SKIP"
    """
    try:
        rating_val = float(rating) if rating and not pd.isna(rating) else 0
    except:
        rating_val = 0
    
    if severity == 'CRITICAL' and rating_val >= 4.0:
        return "🔥 HOT LEAD"
    elif severity == 'CRITICAL':
        return "🟠 QUALIFIED"
    elif severity in ['HIGH', 'MEDIUM']:
        return "🟡 POTENTIAL"
    else:
        return "⚪ SKIP"



def get_technical_notes_severity(notes: str) -> str:
    """
    Анализирует Technical Notes и возвращает severity level.
    
    Returns:
        'CRITICAL' | 'HIGH' | 'MEDIUM' | 'OK'
    """
    if not notes or notes.strip() == "—":
        return 'OK'
    
    notes_lower = notes.lower()
    
    # CRITICAL markers
    critical_markers = [
        '❌ http only',
        '⏱️ site unreachable',
        '🔥 http',
        '🎨 adobe flash'
    ]
    
    # HIGH markers
    high_markers = [
        '⚠️ ssl',
        '📱 missing',
        '📱 no responsive',
        '🎨 jquery'
    ]
    
    # MEDIUM markers
    medium_markers = [
        '⚡ slow',
        '⚡ very slow',
        '🎨 table-based',
        '🎨 fixed-width',
        '🎨 outdated',
        '⚠️ outdated'
    ]
    
    # Check in priority order
    for marker in critical_markers:
        if marker in notes_lower:
            return 'CRITICAL'
    
    for marker in high_markers:
        if marker in notes_lower:
            return 'HIGH'
    
    for marker in medium_markers:
        if marker in notes_lower:
            return 'MEDIUM'
    
    if '✅' in notes:
        return 'OK'
    
    return 'MEDIUM'  # default


def apply_excel_formatting(filepath):
    """
    Применяет визуальное форматирование к Excel файлу.
    
    - Цветовое кодирование ячеек
    - Красивые заголовки
    - Границы
    - Автоширина колонок
    """
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Цвета для заливки
    header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
    red_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    orange_fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFBE6", end_color="FFFBE6", fill_type="solid")
    green_fill = PatternFill(start_color="E6F7E6", end_color="E6F7E6", fill_type="solid")
    gray_fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    
    # Шрифты
    header_font = Font(bold=True, color="FFFFFF", size=12)
    bold_font = Font(bold=True)
    
    # Выравнивание
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    top_left_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)  # Для Technical Notes
    
    # Границы
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 1. Форматирование заголовков (первая строка)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    # Закрепляем первую строку (заголовок)
    ws.freeze_panes = 'A2'  # Закрепляет все строки выше A2 (т.е. строку 1)
    
    # 2. Автоматическое растягивание колонок по содержимому
    for col_idx, column in enumerate(ws.columns, 1):
        col_letter = get_column_letter(col_idx)
        
        # Получаем название колонки
        header_cell = ws.cell(1, col_idx)
        column_name = header_cell.value if header_cell.value else ""
        
        # Находим максимальную длину содержимого в колонке
        max_length = 0
        for cell in column:
            try:
                if cell.value:
                    # Конвертируем в строку и измеряем длину
                    cell_value = str(cell.value)
                    cell_length = len(cell_value)
                    
                    # Для эмодзи добавляем дополнительное пространство
                    # (эмодзи занимают ~2 символа визуально)
                    emoji_count = sum(1 for char in cell_value if ord(char) > 0x1F300)
                    cell_length += emoji_count
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Устанавливаем ширину с учетом типа колонки
        # Technical Notes: фиксированная ширина 60 символов
        # Остальные длинные колонки: Category, Name, Address, Website, Email - до 70
        # Остальные - до 50
        if column_name == '💬 Technical Notes':
            adjusted_width = 60  # Фиксированная ширина для Technical Notes
        elif column_name in ['Category', 'Name', 'Address', 'Website', 'Email']:
            max_width = 70
            adjusted_width = max(8, min(max_length + 3, max_width))
        else:
            max_width = 50
            adjusted_width = max(8, min(max_length + 3, max_width))
        
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # 3. Форматирование ячеек данных + цветовое кодирование
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for col_idx, cell in enumerate(row, start=1):
            cell.border = thin_border
            cell.alignment = left_alignment
            
            # Получаем название колонки через заголовок
            column_name = ws.cell(1, col_idx).value
            
            # Technical Notes - специальное форматирование
            if column_name == '💬 Technical Notes':
                cell.alignment = top_left_alignment  # Wrap text + top-left
                # Устанавливаем высоту строки автоматически для переноса текста
                ws.row_dimensions[row_idx].height = None  # Auto-height
                
                # Determine severity and apply color
                notes_text = str(cell.value) if cell.value else ""
                severity = get_technical_notes_severity(notes_text)
                
                if severity == 'CRITICAL':
                    cell.fill = red_fill
                elif severity == 'HIGH':
                    cell.fill = orange_fill
                elif severity == 'MEDIUM':
                    cell.fill = yellow_fill
                elif severity == 'OK':
                    cell.fill = green_fill
                
                continue  # Skip default formatting logic for this cell
            
            # Цветовое кодирование по содержимому
            value = str(cell.value) if cell.value else ""
            
            # Нумерация, Customer Rating и Phone - по центру
            if column_name == '#' or column_name == '⭐ Customer Rating' or column_name == 'Phone':
                cell.alignment = center_alignment
            
            # Website Status - по левому краю с цветом
            if "✅ Online" in value:
                cell.fill = green_fill
            elif "⚠️ Protected" in value:
                cell.fill = orange_fill  # Работает, но блокирует ботов
            elif "❌ Offline" in value or "⏱️ Timeout" in value:
                cell.fill = red_fill
            
            # Security - по левому краю с цветом
            elif "✅ HTTPS + SSL" in value:
                cell.fill = green_fill
            elif "⚠️ HTTPS" in value:
                cell.fill = orange_fill
            elif "❌ HTTP Only" in value:
                cell.fill = red_fill
            
            # Speed Score - по левому краю с цветом
            elif "🟢 Fast" in value:
                cell.fill = green_fill
            elif "🟡 Average" in value:
                cell.fill = yellow_fill
            elif "🔴 Slow" in value:
                cell.fill = red_fill
            
            # Mobile - по левому краю с цветом
            elif "📱 Mobile" in column_name:
                if "✅ Yes" in value:
                    cell.fill = green_fill
                elif "❌ No" in value:
                    cell.fill = red_fill
            
            # Design - по левому краю с цветом
            elif "🎨 Design" in column_name:
                if "✅ Modern" in value:
                    cell.fill = green_fill
                elif "⚠️ Outdated" in value:
                    cell.fill = orange_fill
            
            # SEO - по левому краю с цветом
            elif "🔍 SEO" in column_name:
                if "✅ Optimized" in value:
                    cell.fill = green_fill
                elif "⚠️ Basic" in value:
                    cell.fill = orange_fill
                elif "❌ Missing" in value:
                    cell.fill = red_fill
            
            # Customer Rating - по центру с цветом (зеленый = высокий, красный = низкий)
            elif column_name == '⭐ Customer Rating' and value != "—":
                try:
                    rating_val = float(value)
                    if rating_val >= 4.5:
                        cell.fill = green_fill  # Отлично
                    elif rating_val >= 4.0:
                        cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')  # Светло-зеленый
                    elif rating_val >= 3.5:
                        cell.fill = yellow_fill  # Хорошо
                    elif rating_val >= 3.0:
                        cell.fill = orange_fill  # Средне
                    else:
                        cell.fill = red_fill  # Плохо
                except:
                    pass
            
            # Lead Quality - по левому краю с цветом и жирным
            elif "🔥 HOT LEAD" in value:
                cell.fill = red_fill
                cell.font = bold_font
            elif "🟠 QUALIFIED" in value:
                cell.fill = orange_fill
            elif "🟡 POTENTIAL" in value:
                cell.fill = yellow_fill
            elif "⚪ SKIP" in value:
                cell.fill = gray_fill
    
    # 4. Включаем автофильтры для заголовков
    if ws.max_row > 0:
        ws.auto_filter.ref = ws.dimensions
    
    # Сохранение
    wb.save(filepath)
    print("✅ Применено визуальное форматирование")
    print("✅ Включены автофильтры")


# ============================================================================
# ОСНОВНАЯ ЛОГИКА
# ============================================================================

def main():
    """Основная функция."""
    
    print("=" * 70)
    print("BrokenSite Hunter - Website Audit Script")
    print("=" * 70)
    
    # Проверка аргументов командной строки
    if len(sys.argv) < 2:
        print("\n❌ Ошибка: Не указан input файл!")
        print("\nИспользование:")
        print("  python audit_websites.py input.xlsx [output.xlsx] [--limit N]")
        print("  python audit_websites.py input.csv [output.csv] [--limit N]")
        print("\nПример:")
        print("  python audit_websites.py dentists.xlsx dentists_audited.xlsx")
        print("  python audit_websites.py dentists.xlsx dentists_audited.xlsx --limit 20")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    # Парсинг опционального параметра --limit
    limit = None
    output_file = None
    
    for i in range(2, len(sys.argv)):
        if sys.argv[i] == '--limit' and i + 1 < len(sys.argv):
            try:
                limit = int(sys.argv[i + 1])
            except ValueError:
                print(f"❌ Ошибка: --limit должен быть числом")
                sys.exit(1)
        elif not sys.argv[i].startswith('--') and not sys.argv[i-1] == '--limit':
            output_file = sys.argv[i]
    
    # Определяем формат выходного файла по входному
    if not output_file:
        # Автоматически определяем расширение
        if input_file.endswith('.xlsx'):
            output_file = 'audited_results.xlsx'
        else:
            output_file = 'audited_results.csv'
    
    # Проверка существования файла
    if not os.path.exists(input_file):
        print(f"\n❌ Ошибка: Файл '{input_file}' не найден!")
        sys.exit(1)
    
    # Загрузка данных (поддержка CSV и XLSX)
    print(f"\n📂 Загружаю данные из: {input_file}")
    try:
        if input_file.endswith('.xlsx'):
            # Excel файл
            df = pd.read_excel(input_file, engine='openpyxl')
            print(f"✅ Загружен Excel файл")
        else:
            # CSV файл
            try:
                df = pd.read_csv(input_file, encoding='utf-8')
            except UnicodeDecodeError:
                df = pd.read_csv(input_file, encoding='latin-1')
            print(f"✅ Загружен CSV файл")
    except Exception as e:
        print(f"\n❌ Ошибка при чтении файла: {e}")
        print("\n💡 Установите openpyxl для работы с Excel:")
        print("   pip install openpyxl")
        sys.exit(1)
    
    print(f"✅ Загружено {len(df)} записей")
    print(f"\n📊 Колонки в файле: {', '.join(df.columns.tolist())}")
    

    # ========================================================================
    # ОПРЕДЕЛЕНИЕ КОЛОНКИ С САЙТОМ (ДО ДЕДУПЛИКАЦИИ)
    # ========================================================================
    # Нам нужно знать колонку, чтобы проводить дедупликацию по ней
    
    # Определение колонки с website
    # Приоритет: точное совпадение, затем вхождение подстроки
    website_column = None
    
    # 1. Сначала ищем точное совпадение (Website, website, URL, url)
    exact_matches = ['Website', 'website', 'URL', 'url', 'Site', 'site']
    for col in df.columns:
        if col in exact_matches:
            website_column = col
            break
    
    # 2. Если не нашли точное - ищем частичное совпадение
    if not website_column:
        for col in df.columns:
            col_lower = col.lower()
            if 'website' in col_lower or 'url' in col_lower or 'site' in col_lower:
                website_column = col
                break
    
    # Если все еще не нашли - пропускаем дедупликацию или выходим с ошибкой
    if not website_column:
        print("\n❌ Ошибка: Не найдена колонка с website/URL для дедупликации!")
        print("Доступные колонки:", ', '.join(df.columns.tolist()))
        sys.exit(1)
        
    print(f"✅ Найдена колонка с URL: '{website_column}'")

    # ========================================================================
    # DEDUPLICATION (v2.4)
    # ========================================================================
    df, dedup_log = deduplicate_data(df, website_column)
    

    

    
    # Проверка API ключа
    if PAGESPEED_API_KEY:
        print(f"✅ PageSpeed API ключ найден")
    else:
        print(f"⚠️  PageSpeed API ключ НЕ найден (проверка будет пропущена)")
        print(f"    Установите: set PAGESPEED_API_KEY=your_key")
    
    # Фильтрация: только записи с website
    df_with_websites = df[df[website_column].notna()].copy()
    


    
    # Применяем лимит если указан
    if limit:
        print(f"⚠️  РЕЖИМ ТЕСТА: Проверяю только первые {limit} сайтов")
        df_with_websites = df_with_websites.head(limit)
    
    print(f"\n🔍 Сайтов для проверки: {len(df_with_websites)}")
    
    if len(df_with_websites) == 0:
        print("❌ Нет сайтов для проверки!")
        sys.exit(1)
    
    # AUDIT
    print("\n" + "=" * 70)
    print("🚀 НАЧИНАЮ ПРОВЕРКУ САЙТОВ")
    print("=" * 70)
    
    results = []
    total = len(df_with_websites)
    
    for idx, row in df_with_websites.iterrows():
        url = row[website_column]
        
        # Проверка сайта
        audit_result = audit_website(
            url=url,
            api_key=PAGESPEED_API_KEY,
            index=idx + 1,
            total=total
        )
        
        results.append(audit_result)
        
        # Задержка между запросами
        if idx < total - 1:  # Не ждем после последнего
            time.sleep(DELAY_BETWEEN_REQUESTS)
    
    # Добавление результатов к DataFrame
    print("\n📊 Создаю красивый отчет...")
    
    results_df = pd.DataFrame(results)
    df_with_websites = df_with_websites.reset_index(drop=True)
    
    # Объединение audit результатов с оригинальными данными
    df_combined = pd.concat([df_with_websites, results_df], axis=1)
    
    # ========================================================================
    # СОЗДАНИЕ ФИНАЛЬНОГО КРАСИВОГО ДАТАСЕТА
    # ========================================================================
    
    # Создаем финальный DataFrame в нужном порядке
    df_final = pd.DataFrame()
    
    # ПОРЯДОК КОЛОНОК: # - Category - Name - Address - Website - Phone - Email - Rating ⭐ - 🌐 - 🔒 - ⚡ - 📱 - 🎨 - 🎯
    
    # 0. Нумерация строк (# используется в США)
    df_final['#'] = range(1, len(df_combined) + 1)
    
    # 1. Category (если есть)
    if 'Category' in df_combined.columns:
        df_final['Category'] = df_combined['Category']
    
    # 2. Name (обязательно)
    if 'Name' in df_combined.columns:
        df_final['Name'] = df_combined['Name']
    
    # 3. Address (если есть)
    if 'Address' in df_combined.columns:
        df_final['Address'] = df_combined['Address']
    
    # 4. Website (очищенный URL)
    df_final['Website'] = df_combined[website_column].apply(clean_website_url)
    
    # 5. Phone (если есть)
    if 'Phone' in df_combined.columns:
        df_final['Phone'] = df_combined['Phone']
    
    # 6. Emails (если есть, форматированные)
    # 6. Emails (Smart Merge: Original vs Scraped)
    def get_best_email(row):
        original = row.get('Emails', row.get('Email', ''))
        scraped = row.get('scraped_emails', '')
        
        # Check original validity
        original_str = str(original).strip()
        is_bad_original = (
            not original or 
            pd.isna(original) or 
            'progress' in original_str.lower() or 
            '###' in original_str or 
            original_str.lower() == 'nan'
        )
        
        if not is_bad_original:
            return format_emails(original)
            
        # Fallback to scraped
        scraped_str = str(scraped).strip()
        if scraped and scraped_str and scraped_str.lower() != 'none' and scraped_str.lower() != 'nan':
            return scraped_str
            
        return "—"

    df_final['Email'] = df_combined.apply(get_best_email, axis=1)
    
    # 7. ⭐ Customer Rating (форматированный, БЕЗ звезды в значениях)
    if 'Rating' in df_combined.columns:
        df_final['⭐ Customer Rating'] = df_combined['Rating'].apply(format_rating)
    
    # ========================================================================
    # ДОБАВЛЯЕМ ФОРМАТИРОВАННЫЕ КОЛОНКИ АУДИТА
    # ========================================================================
    
    # 8. Website Status (форматированный)
    df_final['🌐 Website Status'] = df_combined.apply(
        lambda row: format_website_status(
            row['site_available'], 
            row.get('availability_error', None)
        ), 
        axis=1
    )
    
    # 9. Security Status (Updated logic)
    df_final['🔒 Security'] = df_combined.apply(
        lambda row: '❌ Offline' if not row['site_available'] else format_security_status(
            row['is_https'], 
            row['ssl_valid'],
            row.get('ssl_error', None)
        ), 
        axis=1
    )
    
    # 10. Speed Score (Updated logic)
    df_final['⚡ Speed'] = df_combined.apply(
        lambda row: '❌ Offline' if not row['site_available'] else format_speed_score(row['load_time']),
        axis=1
    )
    
    # 11. Mobile Friendly (Updated logic)
    df_final['📱 Mobile'] = df_combined.apply(
        lambda row: 'N/A' if not row['site_available'] else ('✅ Yes' if row['is_mobile_friendly'] else '❌ No'),
        axis=1
    )
    
    # 12. Design Status (Updated logic)
    df_final['🎨 Design'] = df_combined.apply(
        lambda row: 'N/A' if not row['site_available'] else (f"⚠️ Outdated" if row['is_outdated'] else '✅ Modern'),
        axis=1
    )

    # 13. SEO (новая колонка v2.6!)
    df_final['🔍 SEO'] = df_combined['seo_status']
    
    # 13. Technical Notes (v2.3 - НОВАЯ КОЛОНКА!)
    df_final['💬 Technical Notes'] = df_combined['technical_notes']
    
    # 14. Lead Quality (последняя колонка)
    df_final['🎯 Lead Quality'] = df_combined.apply(
        lambda row: determine_lead_quality(
            row['issue_severity'],
            row.get('Rating', 0)
        ),
        axis=1
    )
    
    # ========================================================================
    # СОРТИРОВКА ПО АЛФАВИТУ
    # ========================================================================
    
    # Сортируем по Name если колонка существует
    if 'Name' in df_final.columns:
        df_final = df_final.sort_values('Name', ascending=True)
        df_final = df_final.reset_index(drop=True)
        # Обновляем нумерацию после сортировки
        df_final['#'] = range(1, len(df_final) + 1)
    
    # Сохранение
    print(f"\n💾 Сохраняю результаты в: {output_file}")
    
    if output_file.endswith('.xlsx'):
        # Сохранение в Excel
        df_final.to_excel(output_file, index=False, engine='openpyxl')
        
        # Применяем визуальное форматирование
        print("🎨 Применяю визуальное форматирование...")
        apply_excel_formatting(output_file)
    else:
        # Сохранение в CSV
        df_final.to_csv(output_file, index=False, encoding='utf-8-sig')
    
    # Статистика
    print("\n" + "=" * 70)
    print("📈 СТАТИСТИКА")
    print("=" * 70)
    
    severity_counts = df_combined['issue_severity'].value_counts()
    
    print(f"\n🔴 CRITICAL: {severity_counts.get('CRITICAL', 0)} сайтов")
    print(f"🟠 HIGH:     {severity_counts.get('HIGH', 0)} сайтов")
    print(f"🟡 MEDIUM:   {severity_counts.get('MEDIUM', 0)} сайтов")
    print(f"🟢 LOW:      {severity_counts.get('LOW', 0)} сайтов")
    print(f"✅ OK:       {severity_counts.get('OK', 0)} сайтов")
    
    # Qualified leads (все с проблемами)
    qualified = df_combined[df_combined['issue_severity'].isin(['CRITICAL', 'HIGH', 'MEDIUM'])]
    print(f"\n🎯 QUALIFIED LEADS: {len(qualified)} сайтов (с проблемами)")
    
    # HOT LEADS статистика
    hot_leads = df_final[df_final['🎯 Lead Quality'] == '🔥 HOT LEAD']
    print(f"🔥 HOT LEADS:      {len(hot_leads)} сайтов (CRITICAL + Rating ≥4.0)")
    
    print("\n✅ ГОТОВО!")
    print(f"📄 Результаты сохранены в: {output_file}")
    
    # Рекомендации
    print("\n💡 СЛЕДУЮЩИЕ ШАГИ:")
    print("1. Откройте файл:", output_file)
    print("2. Файл уже отформатирован с цветами и эмодзи!")
    print("3. Отфильтруйте по колонке '🎯 Lead Quality': 🔥 HOT LEAD, 🟠 QUALIFIED")
    print("4. Добавьте email контакты для qualified leads")


if __name__ == '__main__':
    # Отключаем warnings для SSL
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    main()
